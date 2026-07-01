from __future__ import annotations

import json
import os
import urllib.error
import urllib.request
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from dotenv import load_dotenv
from openpyxl import load_workbook

import ecount_client
import poomgo_client
from logger import setup_logger
from transformer import barcode_aliases, load_product_mapping_csv, normalize_code, to_number


KST = timezone(timedelta(hours=9))
PROJECT_ROOT = Path(__file__).resolve().parents[1]
SHEET_ORDER = ["Dashboard", "월재고현황", "하은물류", "논산합본", "네이버출고", "미매핑상품", "SKU마스터", "설정"]


def env_bool(name: str, default: bool = False) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "y"}


def excel_quote(sheet_name: str) -> str:
    escaped = sheet_name.replace("'", "''")
    return f"'{escaped}'"


def cell(value: Any) -> Any:
    return "" if value is None else value


def to_sheet_number(value: Any) -> Any:
    number = to_number(value, default=None)
    if number is None:
        return ""
    if float(number).is_integer():
        return int(number)
    return number


def apply_manual_mapping_rows(sku_barcode_to_code: dict[str, str], mapping_rows: list[dict[str, Any]]) -> int:
    applied = 0
    for row in mapping_rows:
        code = normalize_code(row.get("sheet_code"))
        if not code:
            continue
        for column in ("ecount_code", "poomgo_code", "naver_code"):
            external_code = normalize_code(row.get(column))
            if not external_code:
                continue
            aliases = [external_code, *barcode_aliases(external_code)]
            for alias in dict.fromkeys(aliases):
                if alias:
                    sku_barcode_to_code[alias] = code
                    applied += 1
    return applied


def append_manual_products(products: list[dict[str, Any]], mapping_rows: list[dict[str, Any]]) -> int:
    existing_codes = {normalize_code(product.get("code")) for product in products}
    added = 0
    for row in mapping_rows:
        code = normalize_code(row.get("sheet_code"))
        if not code or code in existing_codes:
            continue
        products.append(
            {
                "number": len(products) + 1,
                "code": code,
                "name": cell(row.get("product_name")),
                "pieces_per_box": "",
                "boxes_per_pallet": "",
                "pieces_per_pallet": "",
            }
        )
        existing_codes.add(code)
        added += 1
    return added


def read_source_workbook() -> tuple[list[dict[str, Any]], list[list[Any]], dict[str, str]]:
    source_path = PROJECT_ROOT / "data" / "processed" / "osp_source_export.xlsx"
    if not source_path.exists():
        raise FileNotFoundError(f"기준 Google Sheet export 파일을 찾지 못했습니다: {source_path}")

    workbook = load_workbook(source_path, data_only=True)
    month = workbook["월재고현황"]
    sku_master = workbook["SKU마스터"]

    products: list[dict[str, Any]] = []
    for row_index in range(7, 400):
        code = normalize_code(month.cell(row=row_index, column=2).value)
        if not code:
            continue
        products.append(
            {
                "number": len(products) + 1,
                "code": code,
                "name": cell(month.cell(row=row_index, column=3).value),
                "pieces_per_box": cell(month.cell(row=row_index, column=18).value),
                "boxes_per_pallet": cell(month.cell(row=row_index, column=19).value),
                "pieces_per_pallet": cell(month.cell(row=row_index, column=20).value),
            }
        )

    sku_values: list[list[Any]] = []
    for row in sku_master.iter_rows(min_row=1, max_row=sku_master.max_row, min_col=1, max_col=15, values_only=True):
        values = [cell(value) for value in row]
        if any(normalize_code(value) for value in values):
            sku_values.append(values)

    mapping_rows = load_product_mapping_csv(PROJECT_ROOT / "config" / "product_mapping.csv")
    append_manual_products(products, mapping_rows)

    sku_barcode_to_code: dict[str, str] = {}
    for row in sku_values[2:]:
        code = normalize_code(row[2] if len(row) > 2 else "")
        if not code:
            continue
        aliases: list[str] = []
        aliases.extend(barcode_aliases(row[0] if len(row) > 0 else ""))
        aliases.extend(barcode_aliases(row[1] if len(row) > 1 else ""))
        for barcode in aliases:
            sku_barcode_to_code.setdefault(barcode, code)

    apply_manual_mapping_rows(sku_barcode_to_code, mapping_rows)

    return products, sku_values, sku_barcode_to_code


def best_effort_refresh_exports(logger: Any) -> None:
    if not env_bool("AUTO_DOWNLOAD_EXPORTS", True):
        logger.info("AUTO_DOWNLOAD_EXPORTS=false 이므로 export 자동 다운로드를 생략합니다.")
        return

    timeout_seconds = int(os.getenv("DOWNLOAD_TIMEOUT_SECONDS", "120"))
    ecount = ecount_client.from_env()
    for warehouse_code in ("N001", "N004"):
        try:
            downloaded = ecount.download_inventory_export(warehouse_code, timeout_seconds=timeout_seconds)
            logger.info("이카운트 %s export 자동 다운로드 성공: %s", warehouse_code, downloaded)
        except Exception as exc:
            logger.warning("이카운트 %s export 자동 다운로드 실패, 기존 파일로 진행: %s", warehouse_code, exc)

    poomgo = poomgo_client.from_env()
    try:
        downloaded = poomgo.download_naver_inventory_export(timeout_seconds=timeout_seconds)
        logger.info("품고/네이버 export 자동 다운로드 성공: %s", downloaded)
    except Exception as exc:
        logger.warning("품고/네이버 export 자동 다운로드 실패, 기존 파일로 진행: %s", exc)


def collect_records(logger: Any) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    ecount_records: list[dict[str, Any]] = []
    ecount = ecount_client.from_env()
    for warehouse_code in ("N001", "N004"):
        records = ecount.collect_inventory(warehouse_code)
        ecount_records.extend(records)
        logger.info("이카운트 %s 조회 성공: SKU %s개", warehouse_code, len(records))

    poomgo = poomgo_client.from_env()
    poomgo_records = poomgo.collect_naver_inventory()
    logger.info("품고/네이버 재고 조회 성공: SKU %s개", len(poomgo_records))
    return ecount_records, poomgo_records


def build_source_rows(
    ecount_records: list[dict[str, Any]],
    poomgo_records: list[dict[str, Any]],
    sku_barcode_to_code: dict[str, str],
    sheet_codes: set[str],
) -> tuple[list[list[Any]], list[list[Any]], list[list[Any]], str]:
    haeun_rows: list[list[Any]] = []
    nonsan_rows: list[list[Any]] = []
    for record in ecount_records:
        code = normalize_code(record.get("product_code"))
        stock = to_sheet_number(record.get("stock_each"))
        warehouse_code = normalize_code(record.get("warehouse_code"))
        if warehouse_code == "N004":
            haeun_rows.append([code, "", record.get("product_name") or "", stock])
        elif warehouse_code == "N001":
            nonsan_rows.append([code, stock, stock])

    naver_rows: list[list[Any]] = []
    collected_at = ""
    for record in poomgo_records:
        sku_no = normalize_code(record.get("product_code") or record.get("seller_product_code"))
        barcode = normalize_code(record.get("external_product_code") or record.get("naver_code") or sku_no)
        candidates = []
        for value in (barcode, sku_no, record.get("seller_product_code")):
            candidates.extend(barcode_aliases(value))
            normalized = normalize_code(value)
            if normalized:
                candidates.append(normalized)

        code = ""
        for candidate in dict.fromkeys(candidates):
            code = sku_barcode_to_code.get(candidate, "")
            if code:
                break
            if candidate in sheet_codes:
                code = candidate
                break

        stock = record.get("available_stock_each")
        if stock in (None, ""):
            stock = record.get("stock_each")
        record_collected_at = normalize_code(record.get("collected_at"))
        collected_at = collected_at or record_collected_at
        naver_rows.append(
            [
                code,
                sku_no,
                barcode,
                to_sheet_number(stock),
                to_sheet_number(record.get("stock_each")),
                record.get("product_name") or "",
                record_collected_at,
            ]
        )

    if not collected_at:
        collected_at = datetime.now(KST).isoformat(timespec="seconds")
    return haeun_rows, nonsan_rows, naver_rows, collected_at


def display_timestamp(value: str) -> str:
    return value.replace("T", " ").replace("+09:00", " KST")


def build_payload_sheets(
    products: list[dict[str, Any]],
    sku_values: list[list[Any]],
    haeun_rows: list[list[Any]],
    nonsan_rows: list[list[Any]],
    naver_rows: list[list[Any]],
    collected_at: str,
) -> dict[str, list[list[Any]]]:
    product_count = len(products)
    last_product_row = product_count + 6
    collected_display = display_timestamp(collected_at)

    haeun_sheet = [
        ["하은물류 자동 수집 데이터", "", "", ""],
        ["품목코드", "", "상품명", "현재고"],
        *haeun_rows,
    ]

    nonsan_sheet = [["품목코드", "현재고", "현재고_복사"], *nonsan_rows]

    naver_sheet = [
        ["품목코드", "SKU번호", "바코드", "가용재고", "총재고", "SKU명", "수집일시"],
        *naver_rows,
    ]

    month_sheet: list[list[Any]] = [
        ["", "", "날짜기준", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", collected_display, "재고합계", "", "", "하은재고", "", "", "논산재고", "", "", "네이버재고", "", "", "", "", "", "", "", ""],
        ["", "", "", "낱개", "박스", "팔렛트", "낱개", "박스", "팔렛트", "낱개", "박스", "팔렛트", "낱개", "박스", "팔렛트", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "현재고", "", "", "현재고", "", "", "현재고", "", "", "현재고", "", "", "최종수집일시", "오류메시지", "", "", "", ""],
        ["번호", "코드", "상품명", "낱개\n재고", "박스\n재고", "팔렛트\n재고", "낱개\n재고", "박스\n재고", "팔렛트\n재고", "낱개\n재고", "박스\n재고", "팔렛트\n재고", "낱개\n재고", "박스\n재고", "팔렛트\n재고", "최종수집일시", "오류메시지", "코드", "낱개/박스", "박스/파렛트", "낱개/파렛트"],
    ]
    for index, product in enumerate(products, start=7):
        month_sheet.append(
            [
                product["number"],
                product["code"],
                product["name"],
                f"=G{index}+J{index}",
                f"=IFERROR(D{index}/S{index},0)",
                f"=IFERROR(D{index}/U{index},0)",
                f"=SUMIF({excel_quote('하은물류')}!$A$3:$A$200,B{index},{excel_quote('하은물류')}!$D$3:$D$200)",
                f"=IFERROR(G{index}/S{index},0)",
                f"=IFERROR(G{index}/U{index},0)",
                f"=SUMIF({excel_quote('논산합본')}!$A$2:$A$200,B{index},{excel_quote('논산합본')}!$B$2:$B$200)",
                f"=IFERROR(J{index}/S{index},0)",
                f"=IFERROR(J{index}/U{index},0)",
                f"=SUMIF({excel_quote('네이버출고')}!$A$2:$A$200,B{index},{excel_quote('네이버출고')}!$D$2:$D$200)",
                f"=IFERROR(M{index}/S{index},0)",
                f"=IFERROR(M{index}/U{index},0)",
                f"=IF(COUNTIF({excel_quote('네이버출고')}!$A$2:$A$200,B{index})>0,$C$2,\"\")",
                f"=IF(COUNTIF({excel_quote('네이버출고')}!$A$2:$A$200,B{index})=0,\"외부 시스템 미조회: naver\",\"\")",
                product["code"],
                product["pieces_per_box"],
                product["boxes_per_pallet"],
                product["pieces_per_pallet"],
            ]
        )
    month_sheet[3][3:6] = [f"=SUM(D7:D{last_product_row})", f"=SUM(E7:E{last_product_row})", f"=SUM(F7:F{last_product_row})"]
    month_sheet[3][6:9] = [f"=SUM(G7:G{last_product_row})", f"=SUM(H7:H{last_product_row})", f"=SUM(I7:I{last_product_row})"]
    month_sheet[3][9:12] = [f"=SUM(J7:J{last_product_row})", f"=SUM(K7:K{last_product_row})", f"=SUM(L7:L{last_product_row})"]
    month_sheet[3][12:15] = [f"=SUM(M7:M{last_product_row})", f"=SUM(N7:N{last_product_row})", f"=SUM(O7:O{last_product_row})"]

    unmapped_rows = [
        [collected_display, "poomgo", "naver", row[2] or row[1], row[5], row[3], "상품 매핑 실패"]
        for row in naver_rows
        if not row[0]
    ]
    unmapped_sheet = [
        ["발견일시", "source", "warehouse_or_channel", "external_code", "external_product_name", "stock_each", "reason"],
        *unmapped_rows,
    ]

    dashboard_sheet: list[list[Any]] = [
        ["OSP 재고 자동 동기화 Dashboard", "", "", "", "", "", "", "", ""],
        ["마지막 성공 업데이트 일시", collected_display, "", "", "", "", "", "", ""],
        ["마지막 실행 결과", "성공", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["지표", "값", "", "", "", "", "", "", ""],
        ["총 재고합계 낱개", f"=SUM({excel_quote('월재고현황')}!D7:D{last_product_row})", "", "", "", "", "", "", ""],
        ["총 하은재고 낱개", f"=SUM({excel_quote('월재고현황')}!G7:G{last_product_row})", "", "", "", "", "", "", ""],
        ["총 논산재고 낱개", f"=SUM({excel_quote('월재고현황')}!J7:J{last_product_row})", "", "", "", "", "", "", ""],
        ["총 네이버재고 낱개", f"=SUM({excel_quote('월재고현황')}!M7:M{last_product_row})", "", "", "", "", "", "", ""],
        ["품절 SKU 수", f"=COUNTIF({excel_quote('월재고현황')}!D7:D{last_product_row},\"<=0\")", "", "", "", "", "", "", ""],
        ["30개 이하 SKU 수", f"=COUNTIFS({excel_quote('월재고현황')}!D7:D{last_product_row},\">0\",{excel_quote('월재고현황')}!D7:D{last_product_row},\"<=30\")", "", "", "", "", "", "", ""],
        ["네이버 품절 SKU 수", f"=COUNTIF({excel_quote('월재고현황')}!M7:M{last_product_row},\"<=0\")", "", "", "", "", "", "", ""],
        ["미매핑 SKU 수", "=COUNTA('미매핑상품'!D2:D200)", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", ""],
        ["코드", "상품명", "재고합계_낱개", "하은재고_낱개", "논산재고_낱개", "네이버재고_낱개", "전체재고상태", "네이버재고상태", "오류메시지"],
    ]
    for index, product in enumerate(products, start=17):
        month_row = index - 10
        dashboard_sheet.append(
            [
                f"={excel_quote('월재고현황')}!B{month_row}",
                f"={excel_quote('월재고현황')}!C{month_row}",
                f"={excel_quote('월재고현황')}!D{month_row}",
                f"={excel_quote('월재고현황')}!G{month_row}",
                f"={excel_quote('월재고현황')}!J{month_row}",
                f"={excel_quote('월재고현황')}!M{month_row}",
                f"=IF(C{index}<=0,\"품절\",IF(C{index}<=30,\"주의\",\"정상\"))",
                f"=IF(F{index}<=0,\"네이버품절\",IF(F{index}<=30,\"네이버주의\",\"네이버정상\"))",
                f"={excel_quote('월재고현황')}!Q{month_row}",
            ]
        )

    settings_sheet = [
        ["항목", "값"],
        ["문서 용도", "OSP 재고 하루 1회 자동 동기화 대시보드"],
        ["생성 기준", collected_display],
        ["이카운트 하은 SKU 수", len(haeun_rows)],
        ["이카운트 논산 SKU 수", len(nonsan_rows)],
        ["품고/네이버 SKU 수", len(naver_rows)],
        ["월재고현황 SKU 수", len(products)],
        ["비고", "Apps Script 웹 앱으로 자동 갱신됩니다."],
    ]

    return {
        "Dashboard": dashboard_sheet,
        "월재고현황": month_sheet,
        "하은물류": haeun_sheet,
        "논산합본": nonsan_sheet,
        "네이버출고": naver_sheet,
        "미매핑상품": unmapped_sheet,
        "SKU마스터": sku_values,
        "설정": settings_sheet,
    }


def save_payload_snapshot(sheets: dict[str, list[list[Any]]], stats: dict[str, Any]) -> None:
    output_path = PROJECT_ROOT / "data" / "processed" / "apps_script_payload.json"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps({"stats": stats, "sheets": sheets}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def post_to_apps_script(sheets: dict[str, list[list[Any]]]) -> dict[str, Any]:
    url = normalize_code(os.getenv("APPS_SCRIPT_WEBAPP_URL") or os.getenv("GOOGLE_APPS_SCRIPT_URL"))
    token = normalize_code(os.getenv("APPS_SCRIPT_TOKEN", "osp-sync-642768"))
    if not url:
        raise ValueError("APPS_SCRIPT_WEBAPP_URL이 .env에 없습니다.")

    body = json.dumps({"token": token, "sheets": {name: sheets[name] for name in SHEET_ORDER}}, ensure_ascii=False).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json; charset=utf-8"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            text = response.read().decode("utf-8")
    except urllib.error.HTTPError as exc:
        text = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"Apps Script HTTP 오류 {exc.code}: {text}") from exc

    try:
        result = json.loads(text)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Apps Script 응답을 JSON으로 해석하지 못했습니다: {text[:500]}") from exc
    if not result.get("ok"):
        raise RuntimeError(f"Apps Script 동기화 실패: {result}")
    return result


def run_sync() -> int:
    load_dotenv(PROJECT_ROOT / ".env")
    logger = setup_logger(PROJECT_ROOT)
    started_at = datetime.now(KST)
    logger.info("Apps Script 동기화 시작: %s", started_at.isoformat(timespec="seconds"))
    try:
        best_effort_refresh_exports(logger)
        products, sku_values, sku_barcode_to_code = read_source_workbook()
        sheet_codes = {product["code"] for product in products}
        ecount_records, poomgo_records = collect_records(logger)
        haeun_rows, nonsan_rows, naver_rows, collected_at = build_source_rows(
            ecount_records=ecount_records,
            poomgo_records=poomgo_records,
            sku_barcode_to_code=sku_barcode_to_code,
            sheet_codes=sheet_codes,
        )
        sheets = build_payload_sheets(
            products=products,
            sku_values=sku_values,
            haeun_rows=haeun_rows,
            nonsan_rows=nonsan_rows,
            naver_rows=naver_rows,
            collected_at=collected_at,
        )
        stats = {
            "started_at": started_at.isoformat(timespec="seconds"),
            "collected_at": collected_at,
            "n001": len(nonsan_rows),
            "n004": len(haeun_rows),
            "poomgo": len(naver_rows),
            "unmapped": max(len(sheets["미매핑상품"]) - 1, 0),
        }
        save_payload_snapshot(sheets, stats)
        result = post_to_apps_script(sheets)
        finished_at = datetime.now(KST)
        stats["finished_at"] = finished_at.isoformat(timespec="seconds")
        stats["result"] = result
        (PROJECT_ROOT / "data" / "processed" / "last_result.json").write_text(
            json.dumps(stats, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        logger.info("Apps Script 동기화 완료: %s", result)
        return 0
    except Exception as exc:
        logger.exception("Apps Script 동기화 실패: %s", exc)
        return 1


if __name__ == "__main__":
    raise SystemExit(run_sync())
