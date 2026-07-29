from __future__ import annotations

import html
import math
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


OUT_DIR = Path(r"C:\Users\Administrator\Documents\New project\outputs\petfriends_operation_strategy_20260729")
CONVERSION_XLSX = Path(r"C:\Users\CHOIIH\Downloads\(다운로드) 상품별 구간별 전환율 (표).xlsx")
RETENTION_XLSX = Path(r"C:\Users\CHOIIH\Downloads\(다운로드) 재구매율(%) 지표 표.xlsx")
OUTPUT_HTML = OUT_DIR / "osp_petfriends_operation_strategy_20260729.html"


SALES_ROWS = [
    ("하이포-캣", 322500, 15, 408500, 19, 387000, 18, 428464, 20),
    ("하이포-독", 430000, 20, 172000, 8, 344000, 16, 380857, 18),
    ("에이지-캣", 388000, 20, 515000, 27, 709000, 37, 784964, 41),
    ("에이지-독", 18000, 1, 36000, 2, 36000, 2, 39857, 2),
    ("바이오-캣", 216000, 8, 567000, 21, 675000, 25, 747321, 28),
    ("바이오-독", 81000, 3, 108000, 4, 270000, 10, 298929, 11),
    ("세븐오리진-캣", 92500, 5, 240500, 13, 222000, 12, 245786, 13),
    ("세븐오리진-독", 20500, 1, 102500, 5, 143500, 7, 158875, 8),
    ("간식캔", 346500, 231, 688500, 459, 789000, 526, 873536, 582),
]


SKU_SALES_ROWS = [
    ("인디고 독 면역 앤 하이포알러제닉 1kg", "하이포-독", 430000, 20, 172000, 8, 344000, 16, 380857, 18),
    ("인디고 캣 면역 앤 장 키튼 2kg", "바이오-캣", 81000, 3, 135000, 5, 162000, 6, 179357, 7),
    ("인디고 캣 면역 앤 유리너리 2kg", "바이오-캣", 81000, 3, 135000, 5, 324000, 12, 358714, 13),
    ("인디고 캣 면역 앤 헤어볼 2kg", "바이오-캣", 54000, 2, 297000, 11, 189000, 7, 209250, 8),
    ("인디고 캣 면역 앤 하이포알러제닉 1kg", "하이포-캣", 322500, 15, 408500, 19, 387000, 18, 428464, 20),
    ("인디고 에이지뉴트리션 캣 어덜트 1.2kg", "에이지-캣", 36000, 2, 108000, 6, 180000, 10, 199286, 11),
    ("인디고 에이지뉴트리션 캣 어덜트 7+ 1.2kg", "에이지-캣", 152000, 8, 247000, 13, 209000, 11, 231393, 12),
    ("인디고 에이지뉴트리션 캣 어덜트 10+ 1.2kg", "에이지-캣", 200000, 10, 160000, 8, 320000, 16, 354286, 18),
    ("인디고 캣 흰살참치와 게맛살 그레이비캔 80g", "간식캔", 70500, 47, 121500, 81, 118500, 79, 131196, 87),
    ("인디고 캣 흰살참치와 새우 그레이비캔 80g", "간식캔", 45000, 30, 96000, 64, 109500, 73, 121232, 81),
    ("인디고 캣 흰살참치와 멸치치어 그레이비캔 80g", "간식캔", 40500, 27, 87000, 58, 99000, 66, 109607, 73),
    ("인디고 캣 흰살참치와 연어 그레이비캔 80g", "간식캔", 37500, 25, 93000, 62, 121500, 81, 134518, 90),
    ("인디고 캣 흰살참치와 치킨 그레이비캔 80g", "간식캔", 82500, 55, 169500, 113, 208500, 139, 230839, 154),
    ("인디고 캣 흰살참치 그레이비캔 80g", "간식캔", 70500, 47, 121500, 81, 132000, 88, 146143, 97),
    ("인디고 세븐 캣 오리진 로우팻 1.2kg", "세븐오리진-캣", 74000, 4, 148000, 8, 92500, 5, 102411, 6),
    ("인디고 세븐 캣 오리진 하이프로틴 1.2kg", "세븐오리진-캣", 18500, 1, 92500, 5, 129500, 7, 143375, 8),
    ("인디고 세븐 독 오리진 연어&치킨 1.4kg", "세븐오리진-독", 20500, 1, 0, 0, 102500, 5, 113482, 6),
    ("인디고 세븐 독 오리진 소고기&오리 1.4kg", "세븐오리진-독", 0, 0, 102500, 5, 41000, 2, 45393, 2),
    ("인디고 에이지뉴트리션 독 어덜트 1.2kg", "에이지-독", 18000, 1, 36000, 2, 36000, 2, 39857, 2),
    ("인디고 독 면역 앤 체중 2kg", "바이오-독", 54000, 2, 81000, 3, 27000, 1, 29893, 1),
    ("인디고 독 면역 앤 관절 2kg", "바이오-독", 0, 0, 27000, 1, 108000, 4, 119571, 4),
    ("인디고 독 면역 앤 스킨 2kg", "바이오-독", 27000, 1, 0, 0, 135000, 5, 149464, 6),
]


INVENTORY_ROWS = [
    ("하이포-독", "인디고 독 면역 앤 하이포알러제닉 1kg", 16, 32, 33.9, 36, 33),
    ("바이오-캣", "인디고 캣 면역 앤 장 키튼 2kg", 6, 0, 10.2, 11, 33),
    ("바이오-캣", "인디고 캣 면역 앤 유리너리 2kg", 12, 10, 12.4, 15, 37),
    ("바이오-캣", "인디고 캣 면역 앤 헤어볼 2kg", 7, 0, 15.6, 11, 22),
    ("하이포-캣", "인디고 캣 면역 앤 하이포알러제닉 1kg", 18, 8, 41.6, 22, 16),
    ("에이지-캣", "인디고 에이지뉴트리션 캣 어덜트 1.2kg", 10, 0, 11.7, 20, 53),
    ("에이지-캣", "인디고 에이지뉴트리션 캣 어덜트 7+ 1.2kg", 11, 16, 25.1, 29, 36),
    ("에이지-캣", "인디고 에이지뉴트리션 캣 어덜트 10+ 1.2kg", 16, 8, 23.9, 14, 18),
    ("간식캔", "인디고 캣 흰살참치와 게맛살 그레이비캔 80g", 79, 48, 157.2, 90, 18),
    ("간식캔", "인디고 캣 흰살참치와 새우 그레이비캔 80g", 73, 48, 120.9, 113, 29),
    ("간식캔", "인디고 캣 흰살참치와 멸치치어 그레이비캔 80g", 66, 0, 109.4, 51, 14),
    ("간식캔", "인디고 캣 흰살참치와 연어 그레이비캔 80g", 81, 96, 116.9, 158, 42),
    ("간식캔", "인디고 캣 흰살참치와 치킨 그레이비캔 80g", 139, 48, 219.3, 70, 10),
    ("간식캔", "인디고 캣 흰살참치 그레이비캔 80g", 88, 120, 160.5, 177, 34),
    ("세븐오리진-캣", "인디고 세븐 캣 오리진 로우팻 1.2kg", 5, 0, 13.8, 15, 34),
    ("세븐오리진-캣", "인디고 세븐 캣 오리진 하이프로틴 1.2kg", 7, 0, 8.6, 16, 58),
    ("세븐오리진-독", "인디고 세븐 독 오리진 소고기&오리 1.4kg", 2, 0, 5.7, 10, 54),
    ("세븐오리진-독", "인디고 세븐 독 오리진 연어&치킨 1.4kg", 5, 0, 2.8, 6, 65),
    ("에이지-독", "인디고 에이지뉴트리션 독 어덜트 1.2kg", 2, 0, 3.7, 13, 108),
    ("바이오-독", "인디고 독 면역 앤 체중 2kg", 1, 0, 5.4, 7, 40),
    ("바이오-독", "인디고 독 면역 앤 관절 2kg", 4, 0, 2.5, 6, 75),
    ("바이오-독", "인디고 독 면역 앤 스킨 2kg", 5, 0, 2.8, 4, 44),
]


def esc(value: object) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def money(value: float | int) -> str:
    return f"{int(round(value)):,}원"


def manwon(value: float | int) -> str:
    return f"{value / 10000:,.1f}만원"


def pct(value: float | int | None, digits: int = 1) -> str:
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return "-"
    return f"{value * 100:.{digits}f}%"


def delta(new: float, old: float) -> float | None:
    if not old:
        return None
    return new / old - 1


def to_int(value: object) -> int:
    if value in (None, ""):
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    return int(str(value).replace(",", "").strip())


def to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if text.endswith("%"):
        return float(text[:-1]) / 100
    return float(text)


def line_of(name: str) -> str:
    if "그레이비캔" in name:
        return "간식캔"
    if "면역 앤 하이포알러제닉" in name:
        return "하이포"
    if "면역 앤" in name:
        return "바이오"
    if "세븐" in name:
        return "세븐오리진"
    if "에이지뉴트리션" in name:
        return "에이지뉴트리션"
    if "사료샘플" in name:
        return "샘플"
    return "기타"


def pet_of_line(line: str) -> str:
    if line.endswith("-캣") or line == "간식캔":
        return "고양이"
    if line.endswith("-독"):
        return "강아지"
    return "-"


def load_conversion():
    wb = openpyxl.load_workbook(CONVERSION_XLSX, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    total = None
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, raw))
        name = row.get("상품명")
        if not name:
            continue
        if name == "총합계":
            total = row
            continue
        rows.append(row)
    return total, rows


def load_retention():
    wb = openpyxl.load_workbook(RETENTION_XLSX, data_only=True)
    ws = wb.active
    rows = []
    last = {}
    for raw in ws.iter_rows(min_row=4, values_only=True):
        values = list(raw[:13])
        if not any(values):
            continue
        for idx, key in enumerate(["브랜드", "상품명", "1차 카테고리", "2차 카테고리", "대표반려동물"]):
            if values[idx] not in (None, ""):
                last[key] = values[idx]
            else:
                values[idx] = last.get(key)
        if not values[1] or not values[5]:
            continue
        rows.append(
            {
                "name": values[1],
                "line": line_of(values[1]),
                "pet": values[2],
                "month": values[5],
                "first": to_int(values[6]),
                "m1": to_int(values[7]),
                "m1_rate": to_float(values[8]),
                "m2": to_int(values[9]),
                "m2_rate": to_float(values[10]),
            }
        )
    return rows


def conversion_summaries(conv_rows):
    by_cat = defaultdict(lambda: {"views": 0, "cart": 0, "orders": 0})
    by_line = defaultdict(lambda: {"views": 0, "cart": 0, "orders": 0})
    for row in conv_rows:
        cat = f"{row.get('1차 카테고리')}/{row.get('2차 카테고리')}"
        line = line_of(row["상품명"])
        for bucket in (by_cat[cat], by_line[line]):
            bucket["views"] += to_int(row.get("상품상세보기"))
            bucket["cart"] += to_int(row.get("장바구니담기"))
            bucket["orders"] += to_int(row.get("주문완료"))
    for mapping in (by_cat, by_line):
        for value in mapping.values():
            value["cart_rate"] = value["cart"] / value["views"] if value["views"] else None
            value["order_rate"] = value["orders"] / value["views"] if value["views"] else None
    return by_cat, by_line


def retention_summaries(ret_rows):
    by_line = defaultdict(lambda: {"first": 0, "m1": 0, "m2_first": 0, "m2": 0})
    by_pet = defaultdict(lambda: {"first": 0, "m1": 0, "m2_first": 0, "m2": 0})
    for row in ret_rows:
        for bucket in (by_line[row["line"]], by_pet[row["pet"]]):
            bucket["first"] += row["first"]
            bucket["m1"] += row["m1"]
            if row["m2_rate"] is not None:
                bucket["m2_first"] += row["first"]
                bucket["m2"] += row["m2"]
    for mapping in (by_line, by_pet):
        for bucket in mapping.values():
            bucket["m1_rate"] = bucket["m1"] / bucket["first"] if bucket["first"] else None
            bucket["m2_rate"] = bucket["m2"] / bucket["m2_first"] if bucket["m2_first"] else None
    return by_line, by_pet


def tr(cells: list[str], cls: str = "") -> str:
    class_attr = f' class="{cls}"' if cls else ""
    return f"<tr{class_attr}>" + "".join(cells) + "</tr>"


def td(value: object, cls: str = "") -> str:
    class_attr = f' class="{cls}"' if cls else ""
    return f"<td{class_attr}>{esc(value)}</td>"


def th(value: object, cls: str = "") -> str:
    class_attr = f' class="{cls}"' if cls else ""
    return f"<th{class_attr}>{esc(value)}</th>"


def make_bar(label: str, value: float, max_value: float, klass: str = "") -> str:
    height = max(8, round(value / max_value * 100, 1)) if max_value else 8
    return f"""
            <div class="bar-item">
              <div class="bar-value">{esc(manwon(value))}</div>
              <div class="bar-column {klass}" style="height:{height}%"></div>
              <div class="bar-label">{esc(label)}</div>
            </div>"""


def build_report():
    total_conv, conv_rows = load_conversion()
    ret_rows = load_retention()
    conv_by_cat, conv_by_line = conversion_summaries(conv_rows)
    ret_by_line, ret_by_pet = retention_summaries(ret_rows)

    may = sum(row[1] for row in SALES_ROWS)
    june = sum(row[3] for row in SALES_ROWS)
    july = sum(row[5] for row in SALES_ROWS)
    july_expected = sum(row[7] for row in SALES_ROWS)
    max_sales = max(may, june, july, july_expected)

    total_views = to_int(total_conv["상품상세보기"])
    total_orders = to_int(total_conv["주문완료"])
    total_cart_rate = to_float(total_conv["장바구니 전환율"])
    total_order_rate = to_float(total_conv["주문 전환율"])

    total_first = sum(row["first"] for row in ret_rows)
    total_m1 = sum(row["m1"] for row in ret_rows)
    total_m1_rate = total_m1 / total_first if total_first else None

    cat_rows = []
    for name, may_m, may_q, june_m, june_q, july_m, july_q, exp_m, exp_q in SALES_ROWS:
        growth = delta(exp_m, june_m)
        mix = exp_m / july_expected if july_expected else 0
        focus = "성장 확대" if growth and growth >= 0.25 else "방어/검증"
        if name == "간식캔":
            focus = "전환 엔진"
        elif name.endswith("-독") and exp_m < 200000:
            focus = "선별 학습"
        cat_rows.append(
            tr(
                [
                    td(name),
                    td(manwon(may_m), "num"),
                    td(manwon(june_m), "num"),
                    td(manwon(july_m), "num"),
                    td(manwon(exp_m), "num"),
                    td(pct(growth), "num"),
                    td(pct(mix), "num"),
                    td(focus),
                ],
                "focus-row" if name in ("간식캔", "에이지-캣", "바이오-캣") else "",
            )
        )

    top_sku_rows = []
    for row in sorted(SKU_SALES_ROWS, key=lambda x: x[8], reverse=True)[:10]:
        name, line, may_m, may_q, june_m, june_q, july_m, july_q, exp_m, exp_q = row
        top_sku_rows.append(
            tr(
                [
                    td(name),
                    td(line),
                    td(manwon(june_m), "num"),
                    td(manwon(july_m), "num"),
                    td(manwon(exp_m), "num"),
                    td(pct(delta(exp_m, june_m), 1), "num"),
                    td(exp_q, "num"),
                ]
            )
        )

    conv_cat_rows = []
    for name, stats in sorted(conv_by_cat.items(), key=lambda item: item[1]["orders"], reverse=True):
        conv_cat_rows.append(
            tr(
                [
                    td(name),
                    td(f"{stats['views']:,}", "num"),
                    td(pct(stats["cart_rate"]), "num"),
                    td(f"{stats['cart']:,}", "num"),
                    td(pct(stats["order_rate"]), "num"),
                    td(f"{stats['orders']:,}", "num"),
                    td(pct(stats["orders"] / total_orders if total_orders else 0), "num"),
                ],
                "focus-row" if name == "고양이/간식" else "",
            )
        )

    conv_line_rows = []
    for name, stats in sorted(conv_by_line.items(), key=lambda item: item[1]["order_rate"], reverse=True):
        conv_line_rows.append(
            tr(
                [
                    td(name),
                    td(f"{stats['views']:,}", "num"),
                    td(pct(stats["cart_rate"]), "num"),
                    td(pct(stats["order_rate"]), "num"),
                    td(f"{stats['orders']:,}", "num"),
                    td(
                        "확대" if stats["order_rate"] and stats["order_rate"] >= 0.18 else "보완"
                        if stats["order_rate"] and stats["order_rate"] >= 0.09
                        else "재설계"
                    ),
                ]
            )
        )

    top_order_rows = []
    for row in sorted(conv_rows, key=lambda item: to_int(item.get("주문완료")), reverse=True)[:10]:
        top_order_rows.append(
            tr(
                [
                    td(row["상품명"]),
                    td(f"{to_int(row.get('상품상세보기')):,}", "num"),
                    td(pct(to_float(row.get("장바구니 전환율"))), "num"),
                    td(pct(to_float(row.get("주문 전환율"))), "num"),
                    td(f"{to_int(row.get('주문완료')):,}", "num"),
                ],
                "focus-row" if "그레이비캔" in row["상품명"] else "",
            )
        )

    low_conv_rows = []
    low_conv_candidates = [
        row
        for row in conv_rows
        if to_int(row.get("상품상세보기")) >= 100 and to_float(row.get("주문 전환율")) is not None
    ]
    for row in sorted(low_conv_candidates, key=lambda item: to_float(item.get("주문 전환율")))[:8]:
        low_conv_rows.append(
            tr(
                [
                    td(row["상품명"]),
                    td(f"{to_int(row.get('상품상세보기')):,}", "num"),
                    td(pct(to_float(row.get("장바구니 전환율"))), "num"),
                    td(pct(to_float(row.get("주문 전환율"))), "num"),
                    td(f"{to_int(row.get('주문완료')):,}", "num"),
                    td("상세/가격/리뷰 보완"),
                ]
            )
        )

    ret_line_rows = []
    for name, stats in sorted(ret_by_line.items(), key=lambda item: item[1]["m1_rate"] or 0, reverse=True):
        ret_line_rows.append(
            tr(
                [
                    td(name),
                    td(f"{stats['first']:,}", "num"),
                    td(f"{stats['m1']:,}", "num"),
                    td(pct(stats["m1_rate"]), "num"),
                    td(f"{stats['m2']:,}", "num"),
                    td(pct(stats["m2_rate"]), "num"),
                    td(
                        "재구매 앵커" if stats["m1_rate"] and stats["m1_rate"] >= 0.23 else "CRM 유지"
                        if stats["m1_rate"] and stats["m1_rate"] >= 0.17
                        else "재구매 보완"
                    ),
                ],
                "focus-row" if name in ("하이포", "바이오", "간식캔") else "",
            )
        )

    ret_top_rows = []
    for row in sorted([r for r in ret_rows if r["first"] >= 3], key=lambda x: (x["m1_rate"] or 0, x["first"]), reverse=True)[:12]:
        ret_top_rows.append(
            tr(
                [
                    td(row["name"]),
                    td(row["month"]),
                    td(row["first"], "num"),
                    td(row["m1"], "num"),
                    td(pct(row["m1_rate"]), "num"),
                    td(row["m2"], "num"),
                    td(pct(row["m2_rate"]), "num"),
                ]
            )
        )

    inv_low_rows = []
    for line, name, sold, supplied, expected, stock, doc in sorted(INVENTORY_ROWS, key=lambda x: x[6])[:8]:
        inv_low_rows.append(
            tr(
                [
                    td(line),
                    td(name),
                    td(sold, "num"),
                    td(supplied, "num"),
                    td(f"{expected:.1f}", "num"),
                    td(stock, "num"),
                    td(doc, "num"),
                    td("발주/입고 우선" if doc <= 18 else "행사 전 재고 확인"),
                ],
                "focus-row" if doc <= 18 else "",
            )
        )

    inv_high_rows = []
    for line, name, sold, supplied, expected, stock, doc in sorted(INVENTORY_ROWS, key=lambda x: x[6], reverse=True)[:8]:
        inv_high_rows.append(
            tr(
                [
                    td(line),
                    td(name),
                    td(sold, "num"),
                    td(f"{expected:.1f}", "num"),
                    td(stock, "num"),
                    td(doc, "num"),
                    td("프로모션보다 상세/샘플 검증" if sold <= 2 else "소진형 소규모 테스트"),
                ]
            )
        )

    css = """
    :root {
      --ink: #172033;
      --muted: #687386;
      --line: #dbe1e9;
      --bg: #f4f6f8;
      --panel: #ffffff;
      --navy: #23344d;
      --blue: #2563eb;
      --teal: #087a68;
      --green: #147d52;
      --amber: #b46008;
      --red: #b42318;
      --soft-blue: #edf4ff;
      --soft-teal: #eaf7f4;
      --soft-amber: #fff5e6;
      --soft-red: #fff0ee;
    }
    * { box-sizing: border-box; }
    html { scroll-behavior: smooth; }
    body {
      margin: 0;
      color: var(--ink);
      background: var(--bg);
      font-family: "Pretendard", "Apple SD Gothic Neo", "Malgun Gothic", Arial, sans-serif;
      line-height: 1.58;
    }
    a { color: inherit; }
    .shell { width: min(1280px, calc(100% - 40px)); margin: 0 auto; }
    header { border-bottom: 1px solid var(--line); background: var(--panel); }
    header .shell { padding: 30px 0 26px; }
    .badge {
      display: inline-flex; align-items: center; min-height: 26px; padding: 4px 9px;
      border: 1px solid #cbd8ee; border-radius: 5px; color: #2a5598;
      background: var(--soft-blue); font-size: 12px; font-weight: 800;
    }
    h1 { max-width: 900px; margin: 10px 0 12px; color: var(--navy); font-size: 31px; line-height: 1.25; letter-spacing: 0; }
    .meta { display: flex; flex-wrap: wrap; gap: 8px 20px; color: var(--muted); font-size: 13px; }
    .section-nav { position: sticky; top: 0; z-index: 20; border-bottom: 1px solid var(--line); background: rgba(255, 255, 255, 0.96); backdrop-filter: blur(8px); }
    .section-nav .shell { display: flex; gap: 4px; overflow-x: auto; padding: 8px 0; scrollbar-width: thin; }
    .section-nav a { flex: 0 0 auto; padding: 7px 10px; border-radius: 4px; color: #566176; font-size: 13px; font-weight: 700; text-decoration: none; }
    .section-nav a:hover { color: var(--navy); background: #eef2f7; }
    main.shell { padding: 8px 0 72px; }
    section { margin-top: 20px; padding: 24px; border: 1px solid var(--line); border-radius: 7px; background: var(--panel); scroll-margin-top: 58px; }
    section.band { border-color: #cad7e7; background: #f9fbfe; }
    h2 { margin: 0 0 15px; color: var(--navy); font-size: 21px; letter-spacing: 0; }
    h3 { margin: 24px 0 10px; color: var(--navy); font-size: 17px; letter-spacing: 0; }
    p { margin: 8px 0; }
    .lead { max-width: 1000px; color: var(--navy); font-size: 18px; font-weight: 800; }
    .grid { display: grid; gap: 14px; }
    .cols-4 { grid-template-columns: repeat(4, minmax(0, 1fr)); }
    .cols-3 { grid-template-columns: repeat(3, minmax(0, 1fr)); }
    .cols-2 { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    .kpi { min-height: 112px; padding: 15px; border: 1px solid var(--line); border-radius: 6px; background: #fff; }
    .kpi .topline { display: flex; align-items: center; justify-content: space-between; gap: 8px; color: var(--muted); font-size: 12px; font-weight: 700; }
    .kpi .num { margin-top: 12px; color: var(--navy); font-size: 25px; font-weight: 850; line-height: 1.15; }
    .kpi .sub { margin-top: 5px; color: var(--muted); font-size: 12px; }
    .delta { padding: 2px 6px; border-radius: 4px; color: var(--green); background: #e7f6ef; font-size: 11px; font-weight: 800; white-space: nowrap; }
    .delta.down { color: var(--red); background: var(--soft-red); }
    .delta.warn { color: var(--amber); background: var(--soft-amber); }
    .callout { padding: 16px; border: 1px solid var(--line); border-left: 5px solid var(--blue); border-radius: 6px; background: #fbfdff; }
    .callout.green { border-left-color: var(--green); }
    .callout.blue { border-left-color: var(--blue); }
    .callout.amber { border-left-color: var(--amber); }
    .callout.red { border-left-color: var(--red); }
    .callout.teal { border-left-color: var(--teal); }
    .callout .title { color: var(--navy); font-weight: 850; }
    .decision { color: var(--navy); font-weight: 800; }
    .note { margin-top: 10px; color: var(--muted); font-size: 12px; }
    .tag { display: inline-block; padding: 3px 7px; border-radius: 4px; color: #2c55a0; background: var(--soft-blue); font-size: 11px; font-weight: 850; white-space: nowrap; }
    .tag.green { color: var(--green); background: #e8f7ef; }
    .tag.amber { color: var(--amber); background: var(--soft-amber); }
    .tag.red { color: var(--red); background: var(--soft-red); }
    .tag.teal { color: var(--teal); background: var(--soft-teal); }
    .table-wrap { width: 100%; overflow-x: auto; }
    table { width: 100%; margin: 11px 0 5px; border-collapse: collapse; font-size: 13px; }
    th, td { padding: 9px 10px; border: 1px solid var(--line); vertical-align: top; }
    th { color: var(--navy); background: #eef3f8; text-align: left; font-weight: 850; white-space: nowrap; }
    td.num, th.num { text-align: right; white-space: nowrap; }
    tr.focus-row td { background: #fffaf2; }
    .bar-chart {
      display: flex; align-items: flex-end; gap: 14px; height: 230px; padding: 28px 18px 10px;
      border: 1px solid var(--line); border-radius: 6px;
      background:
        linear-gradient(to top, transparent 24%, #edf0f4 25%, transparent 26%),
        linear-gradient(to top, transparent 49%, #edf0f4 50%, transparent 51%),
        linear-gradient(to top, transparent 74%, #edf0f4 75%, transparent 76%),
        #fff;
    }
    .bar-item { display: flex; flex: 1; min-width: 0; height: 100%; flex-direction: column; justify-content: flex-end; align-items: center; }
    .bar-value { margin-bottom: 5px; color: var(--navy); font-size: 13px; font-weight: 850; white-space: nowrap; }
    .bar-column { width: min(72px, 70%); min-height: 8px; border-radius: 4px 4px 0 0; background: var(--blue); }
    .bar-column.teal { background: var(--teal); }
    .bar-column.amber { background: var(--amber); }
    .bar-label { margin-top: 7px; color: var(--muted); font-size: 12px; text-align: center; }
    .meter-row { display: grid; grid-template-columns: 155px 1fr 78px; align-items: center; gap: 10px; margin: 10px 0; font-size: 13px; }
    .meter { height: 11px; overflow: hidden; border-radius: 4px; background: #e8edf3; }
    .meter span { display: block; height: 100%; background: var(--blue); }
    .meter span.teal { background: var(--teal); }
    .meter span.amber { background: var(--amber); }
    .meter span.red { background: var(--red); }
    .ratio-bar { display: flex; height: 34px; overflow: hidden; border-radius: 5px; color: #fff; font-size: 11px; font-weight: 850; }
    .ratio-bar span { display: grid; place-items: center; min-width: 52px; padding: 0 5px; white-space: nowrap; }
    .flow { display: grid; grid-template-columns: 1fr 28px 1fr 28px 1fr 28px 1fr; align-items: stretch; gap: 7px; margin-top: 14px; }
    .flow-step { padding: 14px; border: 1px solid var(--line); border-radius: 5px; background: #fff; }
    .flow-step strong { display: block; margin-bottom: 5px; color: var(--navy); }
    .flow-arrow { display: grid; place-items: center; color: #8994a5; font-weight: 900; }
    .strategy-card { padding: 16px; border-top: 4px solid var(--blue); border-radius: 5px; background: #f9fbfd; }
    .strategy-card.green { border-top-color: var(--green); }
    .strategy-card.amber { border-top-color: var(--amber); }
    .strategy-card.teal { border-top-color: var(--teal); }
    .strategy-card.red { border-top-color: var(--red); }
    .strategy-card h3 { margin: 0 0 8px; font-size: 16px; }
    .strategy-card dl { margin: 10px 0 0; }
    .strategy-card dt { margin-top: 8px; color: var(--muted); font-size: 11px; font-weight: 800; }
    .strategy-card dd { margin: 2px 0 0; font-size: 13px; }
    ul { margin: 8px 0 0 18px; padding: 0; }
    li { margin: 5px 0; }
    .source-list { display: grid; grid-template-columns: 180px 1fr; gap: 6px 14px; color: var(--muted); font-size: 12px; }
    .source-list strong { color: var(--navy); }
    @media (max-width: 980px) {
      .cols-4, .cols-3 { grid-template-columns: repeat(2, minmax(0, 1fr)); }
      .flow { grid-template-columns: 1fr; }
      .flow-arrow { transform: rotate(90deg); min-height: 22px; }
    }
    @media (max-width: 700px) {
      .shell { width: min(100% - 22px, 1280px); }
      header .shell { padding: 23px 0 20px; }
      h1 { font-size: 25px; }
      section { padding: 17px; overflow-x: auto; }
      .cols-4, .cols-3, .cols-2 { grid-template-columns: 1fr; }
      .bar-chart { min-width: 560px; }
      table { min-width: 820px; }
      .meter-row { grid-template-columns: 118px 1fr 62px; }
      .source-list { grid-template-columns: 1fr; }
    }
    @media print {
      body { background: #fff; }
      .section-nav { display: none; }
      .shell { width: 100%; }
      section { break-inside: avoid; border-color: #cfd5dd; }
      .table-wrap { overflow: visible; }
      table { font-size: 10px; }
      .bar-chart { break-inside: avoid; }
    }
    """

    html_doc = f"""<!doctype html>
<html lang="ko">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>펫프렌즈 판매 운영 및 성장 전략 보고서</title>
  <style>{css}</style>
</head>
<body>
  <header>
    <div class="shell">
      <span class="badge">OSP · Pet Friends Operation Intelligence</span>
      <h1>펫프렌즈 판매 운영 및 카테고리 성장 전략 보고서</h1>
      <div class="meta">
        <span>작성일: 2026.07.29</span>
        <span>핵심 실적: 2026.05~2026.07</span>
        <span>7월은 07.28까지 누계, 월말 예상은 일평균 환산</span>
        <span>참고 뷰: 매출 및 Market Share · 재구매율 · 구매 전환율</span>
      </div>
    </div>
  </header>

  <nav class="section-nav">
    <div class="shell">
      <a href="#summary">경영 요약</a>
      <a href="#kpi">지표 체계</a>
      <a href="#sales">매출 흐름</a>
      <a href="#conversion">전환율</a>
      <a href="#retention">재구매</a>
      <a href="#inventory">재고</a>
      <a href="#strategy">상품 전략</a>
      <a href="#playbook">실행안</a>
    </div>
  </nav>

  <main class="shell">
    <section id="summary" class="band">
      <h2>1. 경영 요약</h2>
      <p class="lead">펫프렌즈 운영 방향은 <strong>간식캔으로 전환과 신규 유입을 키우고, 사료는 상품군별 역할을 분리해 재구매로 회수하는 구조</strong>가 맞다. 현재 매출은 성장 중이지만, 강아지 사료와 바이오 일부 SKU는 전환·재구매 검증 없이 광고나 행사를 키우면 효율이 흔들릴 가능성이 높다.</p>
      <div class="grid cols-4" style="margin-top:16px">
        <div class="kpi">
          <div class="topline"><span>7월 월말 예상 매출</span><span class="delta">6월 대비 {pct(delta(july_expected, june))}</span></div>
          <div class="num">{manwon(july_expected)}</div>
          <div class="sub">7월 누계 {manwon(july)} · 예상수량 723개</div>
        </div>
        <div class="kpi">
          <div class="topline"><span>전체 주문 전환율</span><span class="delta">주문 {total_orders:,}건</span></div>
          <div class="num">{pct(total_order_rate)}</div>
          <div class="sub">상세보기 {total_views:,}회 · 장바구니 {pct(total_cart_rate)}</div>
        </div>
        <div class="kpi">
          <div class="topline"><span>고양이 간식 주문 전환율</span><span class="delta">주문 비중 {pct(conv_by_cat['고양이/간식']['orders'] / total_orders)}</span></div>
          <div class="num">{pct(conv_by_cat['고양이/간식']['order_rate'])}</div>
          <div class="sub">상세보기 1,709회 · 주문 514건</div>
        </div>
        <div class="kpi">
          <div class="topline"><span>M+1 재구매율</span><span class="delta warn">초기 코호트</span></div>
          <div class="num">{pct(total_m1_rate)}</div>
          <div class="sub">첫구매 {total_first:,}명 · M+1 재구매 {total_m1:,}명</div>
        </div>
      </div>
      <div class="grid cols-3" style="margin-top:14px">
        <div class="callout green">
          <div class="title">성장 검증</div>
          <p>5월 191.5만원에서 6월 283.8만원, 7월 예상 395.9만원으로 매출이 두 달 연속 성장한다. 에이지-캣, 바이오-캣, 간식캔이 성장을 이끈다.</p>
        </div>
        <div class="callout amber">
          <div class="title">전환 구조</div>
          <p>간식캔은 주문 전환율 30.1%로 전체 평균 18.0%를 크게 상회한다. 반면 강아지 사료는 7.6%로 상품상세와 오퍼 보완이 먼저다.</p>
        </div>
        <div class="callout red">
          <div class="title">재고 조건</div>
          <p>치킨캔 DOC 10일, 멸치치어캔 14일, 캣 하이포 16일은 행사 확대 전 입고가 필요하다. 잘 팔리는 상품일수록 결품이 성장의 병목이 될 수 있다.</p>
        </div>
      </div>
      <p class="decision">최종 판단: 펫프렌즈는 “간식캔 공격 확대 + 고양이 사료 선별 성장 + 강아지 사료 효율 검증”으로 운영한다. 행사는 전환율이 검증된 SKU와 재고가 받쳐주는 SKU부터 적용한다.</p>
    </section>

    <section id="kpi">
      <h2>2. 펫프렌즈 운영 KPI 체계</h2>
      <p class="lead">펫프렌즈 파트너 사이트에서 확인한 핵심 뷰는 매출 및 Market Share, 재구매율, 구매 전환율이다. 여기에 내부 구글시트의 재고회전일과 납품수량을 붙이면 운영 판단이 완성된다.</p>
      <div class="grid cols-4">
        <div class="callout blue">
          <div class="title">1. 매출 및 Market Share</div>
          <p>월별 성장률, 상품군 매출 믹스, 점유율 변화를 보는 상위 지표. 할인보다 먼저 매출 방향성을 확인한다.</p>
        </div>
        <div class="callout green">
          <div class="title">2. 구매 전환율</div>
          <p>상품상세보기, 장바구니, 주문완료를 연결해 상품상세와 가격·리뷰의 힘을 판단한다.</p>
        </div>
        <div class="callout amber">
          <div class="title">3. 재구매율</div>
          <p>M+1, M+2 코호트를 확인해 신규 유입이 단발 구매인지 반복 구매인지 판정한다.</p>
        </div>
        <div class="callout red">
          <div class="title">4. DOC·납품수량</div>
          <p>잘 팔리는 상품이라도 재고가 2~3주 이하라면 행사 확대보다 발주와 입고가 먼저다.</p>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>판단 질문</th>
              <th>볼 지표</th>
              <th>현재 해석</th>
              <th>운영 연결</th>
            </tr>
          </thead>
          <tbody>
            <tr>
              <td>어디가 성장하고 있는가</td>
              <td>월별 매출, Market Share, 상품군 매출 믹스</td>
              <td>7월 예상 매출은 6월 대비 {pct(delta(july_expected, june))} 성장</td>
              <td>에이지-캣, 바이오-캣, 간식캔 중심 확대</td>
            </tr>
            <tr>
              <td>누가 바로 구매하는가</td>
              <td>주문 전환율, 장바구니 전환율</td>
              <td>고양이 간식은 {pct(conv_by_cat['고양이/간식']['order_rate'])}, 강아지 사료는 {pct(conv_by_cat['강아지/사료']['order_rate'])}</td>
              <td>간식캔은 광고·행사 확대, 강아지 사료는 상세 보완 우선</td>
            </tr>
            <tr>
              <td>다시 사는가</td>
              <td>M+1, M+2 재구매율</td>
              <td>하이포 {pct(ret_by_line['하이포']['m1_rate'])}, 바이오 {pct(ret_by_line['바이오']['m1_rate'])}</td>
              <td>하이포는 앵커, 바이오는 CRM 보완</td>
            </tr>
            <tr>
              <td>행사를 받아낼 재고가 있는가</td>
              <td>DOC, 기말재고, 납품수량</td>
              <td>치킨캔 DOC 10일, 캣 하이포 16일</td>
              <td>저 DOC SKU는 입고 후 행사 적용</td>
            </tr>
          </tbody>
        </table>
      </div>
    </section>

    <section id="sales">
      <h2>3. 최근 3개월 판매 흐름</h2>
      <div class="grid cols-2">
        <div>
          <h3 style="margin-top:0">월별 매출 추이</h3>
          <div class="bar-chart" aria-label="펫프렌즈 5월부터 7월 예상 매출 막대그래프">
            {make_bar("5월", may, max_sales)}
            {make_bar("6월", june, max_sales, "teal")}
            {make_bar("7월 누계", july, max_sales, "amber")}
            {make_bar("7월 예상", july_expected, max_sales)}
          </div>
        </div>
        <div>
          <h3 style="margin-top:0">성장 기여 구조</h3>
          <div class="meter-row">
            <span>간식캔 예상 매출 비중</span>
            <div class="meter"><span class="teal" style="width:{round(SALES_ROWS[-1][7] / july_expected * 100, 1)}%"></span></div>
            <strong>{pct(SALES_ROWS[-1][7] / july_expected)}</strong>
          </div>
          <div class="meter-row">
            <span>에이지-캣 예상 비중</span>
            <div class="meter"><span style="width:{round(784964 / july_expected * 100, 1)}%"></span></div>
            <strong>{pct(784964 / july_expected)}</strong>
          </div>
          <div class="meter-row">
            <span>바이오-캣 예상 비중</span>
            <div class="meter"><span style="width:{round(747321 / july_expected * 100, 1)}%"></span></div>
            <strong>{pct(747321 / july_expected)}</strong>
          </div>
          <div class="meter-row">
            <span>강아지 사료 예상 비중</span>
            <div class="meter"><span class="amber" style="width:{round((380857 + 39857 + 298929 + 158875) / july_expected * 100, 1)}%"></span></div>
            <strong>{pct((380857 + 39857 + 298929 + 158875) / july_expected)}</strong>
          </div>
          <div class="callout amber" style="margin-top:18px">
            <div class="title">해석</div>
            <p>매출 성장 자체는 긍정적이다. 다만 7월 성장은 간식캔과 고양이 사료 쪽으로 쏠려 있어, 강아지 사료는 매출 목표보다 먼저 전환율과 재구매를 검증해야 한다.</p>
          </div>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>상품군</th>
              <th class="num">5월 매출</th>
              <th class="num">6월 매출</th>
              <th class="num">7월 누계</th>
              <th class="num">7월 예상</th>
              <th class="num">6월 대비</th>
              <th class="num">예상 믹스</th>
              <th>판단</th>
            </tr>
          </thead>
          <tbody>
            {''.join(cat_rows)}
          </tbody>
        </table>
      </div>

      <h3>7월 예상 매출 상위 SKU</h3>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>상품명</th>
              <th>분류</th>
              <th class="num">6월 매출</th>
              <th class="num">7월 누계</th>
              <th class="num">7월 예상</th>
              <th class="num">6월 대비</th>
              <th class="num">예상수량</th>
            </tr>
          </thead>
          <tbody>
            {''.join(top_sku_rows)}
          </tbody>
        </table>
      </div>
    </section>

    <section id="conversion">
      <h2>4. 상품상세 → 장바구니 → 주문 전환율</h2>
      <p class="lead">전환율 기준으로 보면 펫프렌즈의 현재 승부처는 명확하다. <strong>간식캔은 이미 구매 전환이 검증됐고, 사료는 상품상세·리뷰·가격 제안을 정교하게 다시 잡아야 한다.</strong></p>
      <div class="grid cols-3">
        <div class="callout green">
          <div class="title">고양이 간식</div>
          <p>주문 전환율 {pct(conv_by_cat['고양이/간식']['order_rate'])}, 주문 514건. 전체 주문의 {pct(conv_by_cat['고양이/간식']['orders'] / total_orders)}를 만든다.</p>
        </div>
        <div class="callout amber">
          <div class="title">고양이 사료</div>
          <p>상세보기는 1,667회로 많지만 주문 전환율은 {pct(conv_by_cat['고양이/사료']['order_rate'])}. 질환·연령·기능 키워드별 설득이 필요하다.</p>
        </div>
        <div class="callout red">
          <div class="title">강아지 사료</div>
          <p>주문 전환율 {pct(conv_by_cat['강아지/사료']['order_rate'])}. 광고 증액보다 상세페이지, 리뷰, 체험팩 연결을 먼저 봐야 한다.</p>
        </div>
      </div>
      <h3>카테고리별 전환</h3>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>카테고리</th>
              <th class="num">상품상세보기</th>
              <th class="num">장바구니 전환율</th>
              <th class="num">장바구니담기</th>
              <th class="num">주문 전환율</th>
              <th class="num">주문완료</th>
              <th class="num">주문 비중</th>
            </tr>
          </thead>
          <tbody>
            {''.join(conv_cat_rows)}
          </tbody>
        </table>
      </div>
      <h3>상품군별 전환 판단</h3>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>상품군</th>
              <th class="num">상세보기</th>
              <th class="num">장바구니 전환율</th>
              <th class="num">주문 전환율</th>
              <th class="num">주문</th>
              <th>판정</th>
            </tr>
          </thead>
          <tbody>
            {''.join(conv_line_rows)}
          </tbody>
        </table>
      </div>
      <h3>주문 상위 SKU</h3>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>상품명</th>
              <th class="num">상세보기</th>
              <th class="num">장바구니 전환율</th>
              <th class="num">주문 전환율</th>
              <th class="num">주문</th>
            </tr>
          </thead>
          <tbody>
            {''.join(top_order_rows)}
          </tbody>
        </table>
      </div>
      <h3>전환 보완이 필요한 SKU</h3>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>상품명</th>
              <th class="num">상세보기</th>
              <th class="num">장바구니 전환율</th>
              <th class="num">주문 전환율</th>
              <th class="num">주문</th>
              <th>조치</th>
            </tr>
          </thead>
          <tbody>
            {''.join(low_conv_rows)}
          </tbody>
        </table>
      </div>
    </section>

    <section id="retention">
      <h2>5. 재구매율과 고객 회수 구조</h2>
      <p class="lead">재구매율은 아직 초기 코호트라 표본이 작다. 그래도 방향은 보인다. <strong>하이포와 세븐은 반복 구매 신호가 좋고, 바이오는 매출 성장 대비 M+1 재구매가 약하다.</strong></p>
      <div class="grid cols-3">
        <div class="callout green">
          <div class="title">하이포</div>
          <p>M+1 재구매율 {pct(ret_by_line['하이포']['m1_rate'])}. 특히 독 하이포 1kg은 2026년 5월 첫구매 11명 중 M+1 4명, M+2 3명이다.</p>
        </div>
        <div class="callout amber">
          <div class="title">간식캔</div>
          <p>첫구매 240명으로 신규 유입 규모가 가장 크다. M+1은 {pct(ret_by_line['간식캔']['m1_rate'])}라서 맛별 번들·CRM으로 재구매를 회수해야 한다.</p>
        </div>
        <div class="callout red">
          <div class="title">바이오</div>
          <p>M+1 재구매율 {pct(ret_by_line['바이오']['m1_rate'])}. 매출은 성장 중이나 재구매가 약해 기능별 리마인드와 30일 쿠폰이 필요하다.</p>
        </div>
      </div>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>상품군</th>
              <th class="num">첫구매 고객</th>
              <th class="num">M+1 재구매</th>
              <th class="num">M+1 재구매율</th>
              <th class="num">M+2 재구매</th>
              <th class="num">M+2 재구매율</th>
              <th>판단</th>
            </tr>
          </thead>
          <tbody>
            {''.join(ret_line_rows)}
          </tbody>
        </table>
      </div>
      <h3>재구매 신호가 좋은 코호트</h3>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>상품명</th>
              <th>첫구매월</th>
              <th class="num">첫구매 고객</th>
              <th class="num">M+1 재구매</th>
              <th class="num">M+1 재구매율</th>
              <th class="num">M+2 재구매</th>
              <th class="num">M+2 재구매율</th>
            </tr>
          </thead>
          <tbody>
            {''.join(ret_top_rows)}
          </tbody>
        </table>
      </div>
      <p class="note">재구매율은 2026년 5~6월 첫구매 코호트가 중심이라 아직 표본이 작다. 단일 SKU의 절대 수치보다 상품군별 방향과 M+1 이후 CRM 반응을 함께 봐야 한다.</p>
    </section>

    <section id="inventory">
      <h2>6. 재고와 행사 가능성</h2>
      <p class="lead">펫프렌즈는 전환율 높은 SKU를 키우는 것이 맞지만, <strong>행사 전 재고회전일을 먼저 확인해야 한다.</strong> 특히 간식캔 상위 SKU는 전환이 좋지만 DOC가 낮아 결품 리스크가 있다.</p>
      <div class="grid cols-2">
        <div>
          <h3 style="margin-top:0">발주/입고 우선 SKU</h3>
          <div class="table-wrap">
            <table>
              <thead>
                <tr>
                  <th>분류</th>
                  <th>상품명</th>
                  <th class="num">7월 판매</th>
                  <th class="num">7월 납품</th>
                  <th class="num">예상판매</th>
                  <th class="num">기말재고</th>
                  <th class="num">DOC</th>
                  <th>판단</th>
                </tr>
              </thead>
              <tbody>
                {''.join(inv_low_rows)}
              </tbody>
            </table>
          </div>
        </div>
        <div>
          <h3 style="margin-top:0">소진/검증 우선 SKU</h3>
          <div class="table-wrap">
            <table>
              <thead>
                <tr>
                  <th>분류</th>
                  <th>상품명</th>
                  <th class="num">7월 판매</th>
                  <th class="num">예상판매</th>
                  <th class="num">기말재고</th>
                  <th class="num">DOC</th>
                  <th>판단</th>
                </tr>
              </thead>
              <tbody>
                {''.join(inv_high_rows)}
              </tbody>
            </table>
          </div>
        </div>
      </div>
      <div class="callout amber" style="margin-top:14px">
        <div class="title">운영 기준</div>
        <p>DOC 20일 이하는 행사 확대 전 입고가 먼저다. DOC 50일 이상은 대형 할인보다 상세페이지 개선, 샘플·리뷰 확보, 제한적 쿠폰으로 수요를 검증한다.</p>
      </div>
    </section>

    <section id="strategy">
      <h2>7. 상품군별 운영 전략</h2>
      <div class="grid cols-3">
        <article class="strategy-card green">
          <h3>간식캔</h3>
          <span class="tag green">전환 엔진</span>
          <dl>
            <dt>검증</dt>
            <dd>주문 전환율 {pct(conv_by_line['간식캔']['order_rate'])}, 주문 514건. 전체 주문의 대부분을 만든다.</dd>
            <dt>운영</dt>
            <dd>치킨·게맛살·기본참치를 대표 SKU로 두고, 새우·연어·멸치치어는 맛별 확장과 묶음 구매를 유도한다.</dd>
            <dt>주의</dt>
            <dd>치킨캔 DOC 10일, 멸치치어 14일. 광고/행사 확대 전 입고 계획이 필요하다.</dd>
          </dl>
        </article>
        <article class="strategy-card teal">
          <h3>하이포</h3>
          <span class="tag teal">재구매 앵커</span>
          <dl>
            <dt>검증</dt>
            <dd>M+1 재구매율 {pct(ret_by_line['하이포']['m1_rate'])}. 사료 중 반복 구매 신호가 가장 좋다.</dd>
            <dt>운영</dt>
            <dd>캣 하이포는 재고 보강 후 검색·상세 키워드 강화, 독 하이포는 체험 후 본품 전환 메시지를 유지한다.</dd>
            <dt>KPI</dt>
            <dd>M+1/M+2 재구매, 리뷰 증가, 1kg 구매 후 2kg·대용량 전환.</dd>
          </dl>
        </article>
        <article class="strategy-card amber">
          <h3>바이오</h3>
          <span class="tag amber">성장·재구매 보완</span>
          <dl>
            <dt>검증</dt>
            <dd>바이오-캣 7월 예상은 74.7만원으로 6월 대비 {pct(delta(747321, 567000))} 성장. 단 M+1은 {pct(ret_by_line['바이오']['m1_rate'])}로 낮다.</dd>
            <dt>운영</dt>
            <dd>유리너리는 성장 SKU로 유지하고, 헤어볼·장키튼은 상세페이지와 기능별 CRM을 보완한다.</dd>
            <dt>KPI</dt>
            <dd>상품상세 주문 전환율, 30일 재구매, 기능별 검색 유입.</dd>
          </dl>
        </article>
        <article class="strategy-card amber">
          <h3>에이지뉴트리션</h3>
          <span class="tag amber">고양이 선별 확대</span>
          <dl>
            <dt>검증</dt>
            <dd>에이지-캣 7월 예상 78.5만원, 6월 대비 {pct(delta(784964, 515000))}. 고양이 10+와 7+가 매출 상위권이다.</dd>
            <dt>운영</dt>
            <dd>고양이 10+·7+ 중심으로 연령/시니어 케어 메시지를 강화한다. 독 어덜트는 전환율 4.8%, DOC 108일로 확장 보류.</dd>
            <dt>KPI</dt>
            <dd>연령 키워드 유입, 시니어 고객 신규수, 재구매 전환.</dd>
          </dl>
        </article>
        <article class="strategy-card teal">
          <h3>세븐오리진</h3>
          <span class="tag teal">선별 학습</span>
          <dl>
            <dt>검증</dt>
            <dd>M+1 재구매율은 {pct(ret_by_line['세븐오리진']['m1_rate'])}이나 표본이 작다. 주문 전환율은 {pct(conv_by_line['세븐오리진']['order_rate'])}로 낮다.</dd>
            <dt>운영</dt>
            <dd>로우팻·하이프로틴은 상세와 리뷰를 보완하고, 독 1.4kg은 샘플/체험 중심으로 확인한다.</dd>
            <dt>주의</dt>
            <dd>4.8kg SKU는 납품·재고가 있어도 판매 신호가 약하다. 대용량 확대 전 소포장 반응이 먼저다.</dd>
          </dl>
        </article>
        <article class="strategy-card red">
          <h3>강아지 사료</h3>
          <span class="tag red">효율 검증</span>
          <dl>
            <dt>검증</dt>
            <dd>강아지 사료 주문 전환율은 {pct(conv_by_cat['강아지/사료']['order_rate'])}. 상세보기 대비 주문 전환이 가장 낮다.</dd>
            <dt>운영</dt>
            <dd>광고비 증액보다 상세페이지, 리뷰, 샘플 연결, 첫구매 쿠폰을 먼저 테스트한다.</dd>
            <dt>중단 기준</dt>
            <dd>2주 테스트 후 상세 전환과 장바구니율이 개선되지 않으면 예산 확대를 보류한다.</dd>
          </dl>
        </article>
      </div>
    </section>

    <section id="playbook">
      <h2>8. 향후 90일 실행안</h2>
      <div class="flow">
        <div class="flow-step">
          <strong>1~2주차</strong>
          DOC 20일 이하 SKU 발주 확정, 간식캔 대표 SKU 재고 보강, 강아지 사료 상세·리뷰 점검.
        </div>
        <div class="flow-arrow">→</div>
        <div class="flow-step">
          <strong>3~6주차</strong>
          간식캔 광고·행사 확대, 캣 하이포·유리너리·에이지 10+ 검색 키워드 강화.
        </div>
        <div class="flow-arrow">→</div>
        <div class="flow-step">
          <strong>7~10주차</strong>
          첫구매 D+14, D+30 CRM으로 재구매 쿠폰과 맛별 번들, 사료 전환 메시지 운영.
        </div>
        <div class="flow-arrow">→</div>
        <div class="flow-step">
          <strong>11~12주차</strong>
          SKU별 승격·유지·축소를 결정하고 Market Share와 재구매율을 월간 리포트화.
        </div>
      </div>
      <h3>운영군별 액션</h3>
      <div class="table-wrap">
        <table>
          <thead>
            <tr>
              <th>운영군</th>
              <th>대상</th>
              <th>행사·가격</th>
              <th>광고·노출</th>
              <th>핵심 KPI</th>
              <th>조정 기준</th>
            </tr>
          </thead>
          <tbody>
            <tr>
              <td><span class="tag green">공격형</span></td>
              <td>간식캔 상위 SKU</td>
              <td>입고 확보 후 묶음/맛별 세트 행사</td>
              <td>검색·추천·카테고리 노출 확대</td>
              <td>주문 전환율, 신규고객, DOC</td>
              <td>DOC 20일 이하로 내려가면 행사 강도 조절</td>
            </tr>
            <tr>
              <td><span class="tag teal">앵커형</span></td>
              <td>하이포 캣·독</td>
              <td>상시 고할인보다 재구매 쿠폰 중심</td>
              <td>하이포알러제닉·피부·면역 키워드</td>
              <td>M+1/M+2 재구매율, 리뷰, 재고</td>
              <td>재구매율 유지 시 본품 전환 확대</td>
            </tr>
            <tr>
              <td><span class="tag amber">성장형</span></td>
              <td>유리너리, 에이지 캣 10+·7+</td>
              <td>2주 단위 제한 쿠폰</td>
              <td>기능·연령 키워드와 상세페이지 연결</td>
              <td>상세 전환율, 장바구니율, 재구매</td>
              <td>전환 개선이 없으면 소재와 상세 재정비</td>
            </tr>
            <tr>
              <td><span class="tag red">검증형</span></td>
              <td>강아지 사료, DOC 50일 이상 SKU</td>
              <td>대형 할인 전 소규모 쿠폰·샘플</td>
              <td>체험팩, 리뷰 확보, 문제 해결 메시지</td>
              <td>상세보기 대비 주문, 샘플 후 본품 전환</td>
              <td>2주 후 전환 신호 없으면 광고 확대 보류</td>
            </tr>
          </tbody>
        </table>
      </div>
      <div class="grid cols-2" style="margin-top:16px">
        <div class="callout green">
          <div class="title">지금 바로 할 것</div>
          <ul>
            <li>치킨캔, 멸치치어캔, 캣 하이포, 에이지 캣 10+ 입고 계획 확인</li>
            <li>간식캔 맛별 대표 SKU를 정하고 묶음 구매 노출 강화</li>
            <li>하이포 구매자 D+30 재구매 쿠폰과 본품 전환 CRM 세팅</li>
            <li>유리너리·에이지 10+ 상세페이지의 기능 키워드 정리</li>
          </ul>
        </div>
        <div class="callout red">
          <div class="title">지금 줄일 것</div>
          <ul>
            <li>DOC 20일 이하 SKU의 무리한 행사 확대</li>
            <li>강아지 사료의 전환율 개선 없는 광고 증액</li>
            <li>전 SKU 일괄 할인</li>
            <li>재구매율 확인 없이 신규 유입만 늘리는 운영</li>
          </ul>
        </div>
      </div>
      <p class="lead" style="margin-top:18px">한 줄 결론: 펫프렌즈는 간식캔으로 구매를 만들고, 하이포·유리너리·에이지 고양이 사료로 반복 구매를 회수한다. 강아지 사료는 큰 예산보다 작은 테스트로 전환율을 먼저 끌어올린다.</p>
    </section>

    <section>
      <h2>9. 데이터 출처 및 해석 기준</h2>
      <div class="source-list">
        <strong>펫프렌즈 Tableau</strong>
        <span>[EXPERT] 데이터 솔루션 워크북, 매출 및 Market Share · 재구매율(%) 지표 · 구매 전환율(%) 지표 확인</span>
        <strong>구매 전환율</strong>
        <span>로컬 파일: (다운로드) 상품별 구간별 전환율 (표).xlsx, 상품상세보기·장바구니담기·주문완료 기준</span>
        <strong>재구매율</strong>
        <span>로컬 파일: (다운로드) 재구매율(%) 지표 표.xlsx, 첫구매월별 M+1/M+2 코호트 기준</span>
        <strong>판매·재고</strong>
        <span>Google Sheet: ■ 펫프렌즈 데이터, SKU별 판매현황·SKU별 재고·납품수량·프로모션 탭</span>
        <strong>7월 월말 예상</strong>
        <span>2026.07.29 시트 기준 7월 누계를 월말로 환산. 실제 마감 매출과 달라질 수 있음</span>
        <strong>해석 제한</strong>
        <span>Market Share 값은 Tableau 뷰의 추가 관리 지표로 확인했으며, 보고서 본문 수치 계산은 제공된 엑셀 및 Google Sheet 원천값을 기준으로 작성</span>
      </div>
    </section>
  </main>
</body>
</html>
"""

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_HTML.write_text(html_doc, encoding="utf-8")
    return OUTPUT_HTML


if __name__ == "__main__":
    path = build_report()
    print(path)
